import os
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from django.conf import settings


def generate_filename(user_id):
    # Generate a unique filename based on user_id and timestamp
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"quiz_template_{user_id}_{timestamp}.xlsx"

def insert_questions_into_existing_excel(file_path, num_questions):
    # Load the workbook and select the active sheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Determine how many columns exist (based on header row)
    max_columns = ws.max_column

    # Insert rows after row 2
    ws.insert_rows(idx=2, amount=num_questions)

    # Fill new rows
    for i in range(num_questions):
        row_number = 2 + i  # Start inserting at row 3
        ws.cell(row=row_number, column=1, value=i + 1)  # First column = number

        # Apply borders to all cells in this row
        for col in range(1, max_columns + 1):
            ws.cell(row=row_number, column=col).border = thin_border

    # Save workbook
    wb.save(file_path)
    print(f"✅ {num_questions} questions inserted into '{file_path}' starting from row 3.")

def generate_excel_file(user_id, num_questions):
    # Generate a filename based on user_id and timestamp
    filename = generate_filename(user_id)

    # Ensure the 'media' and 'exports' folders exist
    media_folder = settings.MEDIA_ROOT
    if not os.path.exists(media_folder):
        os.makedirs(media_folder)

    exports_folder = os.path.join(media_folder, 'exports')
    if not os.path.exists(exports_folder):
        os.makedirs(exports_folder)

    # Define the file path where the file will be saved
    file_path = os.path.join(exports_folder, filename)
    
    # Print for debugging purposes
    print(f"Saving file to: {file_path}")

    # Load the template Excel file from the 'templates' directory
    template_path = os.path.join(settings.BASE_DIR, 'apps', 'templates', 'quiz.xlsx')
    print(f"Loading template from: {template_path}")
    
    # Ensure the template exists before proceeding
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template file not found at {template_path}")
    print(f"✅ Template found at '{template_path}'.")
    
    # Create a new workbook with a header using the template
    wb = load_workbook(template_path)
    print(f"✅ Template loaded from '{template_path}'.")
    
    # Save the workbook to the file path before inserting questions
    wb.save(file_path)
    print(f"✅ Workbook saved to '{file_path}'.")

    # Insert questions into the workbook
    try:
        insert_questions_into_existing_excel(file_path, num_questions)
    except Exception as e:
        print(f"❌ An error occurred while inserting questions: {e}")
        raise
    print(f"✅ Excel file '{filename}' generated successfully.")
    
    return filename  # Return the filename for download link
